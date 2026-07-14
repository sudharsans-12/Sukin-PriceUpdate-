"""
Builds the styled Excel output workbook for the Price Update Automation Tool.

Uses openpyxl's declarative conditional formatting (rule-based) rather than
per-cell fill loops, so styling large result sets (100K+ rows) stays fast.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GREY_FONT = Font(color="595959")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(color="9C6500")

OUTPUT_COLUMNS = ["Seller SKU", "RRP", "RRP Check", "SRP", "SRP Check", "Need to Update"]


class ExcelReportBuilder:
    """Builds a styled .xlsx workbook (as bytes) from a validation result DataFrame."""

    def build(self, result_df: pd.DataFrame, sheet_name: str = "Price Validation") -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        self._write_data(sheet, result_df)
        self._style_header(sheet, ncols=len(OUTPUT_COLUMNS))
        self._apply_conditional_formatting(sheet, nrows=len(result_df))
        self._auto_size_columns(sheet, result_df)
        sheet.freeze_panes = "A2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _write_data(sheet: Worksheet, df: pd.DataFrame) -> None:
        sheet.append(OUTPUT_COLUMNS)
        for row in df[OUTPUT_COLUMNS].itertuples(index=False):
            sheet.append(list(row))

    @staticmethod
    def _style_header(sheet: Worksheet, ncols: int) -> None:
        for col_idx in range(1, ncols + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _apply_conditional_formatting(sheet: Worksheet, nrows: int) -> None:
        if nrows == 0:
            return
        last_row = nrows + 1  # +1 for header

        check_ranges = [
            f"C2:C{last_row}",  # RRP Check
            f"E2:E{last_row}",  # SRP Check
        ]
        for cell_range in check_ranges:
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN_FILL, font=GREEN_FONT),
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal", formula=['"Need Update"'], fill=RED_FILL, font=RED_FONT
                ),
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=['"N/A"'], fill=GREY_FILL, font=GREY_FONT),
            )

        need_update_range = f"F2:F{last_row}"
        sheet.conditional_formatting.add(
            need_update_range,
            CellIsRule(operator="equal", formula=['"Yes"'], fill=RED_FILL, font=RED_FONT),
        )
        sheet.conditional_formatting.add(
            need_update_range,
            CellIsRule(operator="equal", formula=['"No"'], fill=GREEN_FILL, font=GREEN_FONT),
        )
        sheet.conditional_formatting.add(
            need_update_range,
            CellIsRule(operator="equal", formula=['"N/A"'], fill=GREY_FILL, font=GREY_FONT),
        )

    @staticmethod
    def _auto_size_columns(sheet: Worksheet, df: pd.DataFrame) -> None:
        for idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            if len(df):
                content_len = int(df[col_name].map(lambda v: len(str(v))).max())
            else:
                content_len = 0
            max_len = max(len(col_name), content_len)
            sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 4, 40)


def build_summary(result_df: pd.DataFrame) -> dict[str, int]:
    """Compute quick summary counts for display in the Streamlit UI."""
    total = len(result_df)
    need_update = int((result_df["Need to Update"] == "Yes").sum())
    ok = int((result_df["Need to Update"] == "No").sum())
    na = int((result_df["Need to Update"] == "N/A").sum())
    return {"Total Rows": total, "Need Update": need_update, "OK": ok, "N/A": na}
